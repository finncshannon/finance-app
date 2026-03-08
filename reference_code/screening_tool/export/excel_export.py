"""
Excel Export -- exports search results to a formatted Excel workbook.

Creates a workbook with:
  Sheet 1: Search Results (ticker, company, match%, sector, financials, model fit)
  Sheet 2: Match Details (matched text excerpts per company)

Usage:
    from export.excel_export import export_results

    path = export_results(results, query, store)
"""

import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime

logger = logging.getLogger(__name__)

EXPORT_DIR = Path(__file__).parent.parent / "data" / "search_results"


def export_results(results: list, query: str, store=None,
                    output_path: Optional[Path] = None) -> str:
    """
    Export search results to Excel.

    Args:
        results: List of SearchResult objects
        query: The search query used
        store: CompanyStore for additional data
        output_path: Custom output path. Default: data/search_results/

    Returns:
        Path to the exported file as string.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_query = "".join(c if c.isalnum() else "_" for c in query[:30])
        filename = f"Screener_{safe_query}_{timestamp}.xlsx"
        output_path = EXPORT_DIR / filename

    wb = Workbook()

    # ---- Sheet 1: Search Results ----
    ws1 = wb.active
    ws1.title = "Search Results"

    # Styles
    header_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0f3460', end_color='0f3460', fill_type='solid')
    data_font = Font(name='Segoe UI', size=9)
    currency_fmt = '#,##0'
    pct_fmt = '0%'

    # Headers
    headers = [
        "Ticker", "Company", "Match %", "Hits", "Sector",
        "Keywords Matched", "Revenue", "Net Income", "FCF",
        "Model Fit", "Recommended", "Filing Date",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, r in enumerate(results, 2):
        model_str = ""
        rec_str = ""
        if hasattr(r, '_model_fit') and r._model_fit:
            model_str = r._model_fit.fit_summary
            rec_str = r._model_fit.recommended

        kw_str = ", ".join(
            f"{kw}({r.keyword_counts.get(kw, 0)})"
            for kw in r.keywords_matched
        )

        row_data = [
            r.ticker,
            r.company_name,
            r.match_score / 100,  # For percentage format
            r.total_hits,
            r.sector,
            kw_str,
            r.revenue,
            r.net_income,
            r.fcf,
            model_str,
            rec_str,
            r.filing_date,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.font = data_font

    # Format columns
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 6
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 35
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 14
    ws1.column_dimensions['I'].width = 14
    ws1.column_dimensions['J'].width = 18
    ws1.column_dimensions['K'].width = 14
    ws1.column_dimensions['L'].width = 12

    # Number formats
    for row in range(2, len(results) + 2):
        ws1.cell(row=row, column=3).number_format = pct_fmt
        for col in [7, 8, 9]:  # Revenue, NI, FCF
            ws1.cell(row=row, column=col).number_format = currency_fmt

    # ---- Sheet 2: Match Details ----
    ws2 = wb.create_sheet("Match Details")

    detail_headers = ["Ticker", "Section", "Excerpt", "Keywords Found"]
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    detail_row = 2
    for r in results:
        section_names = {
            "item1": "Item 1 (Business)",
            "item1a": "Item 1A (Risk Factors)",
            "item7": "Item 7 (MD&A)",
        }

        for excerpt in r.matched_excerpts[:5]:
            ws2.cell(row=detail_row, column=1, value=r.ticker).font = data_font
            ws2.cell(row=detail_row, column=2, value="").font = data_font
            # Truncate excerpts for Excel
            excerpt_clean = excerpt[:500] if excerpt else ""
            ws2.cell(row=detail_row, column=3, value=excerpt_clean).font = data_font
            ws2.cell(row=detail_row, column=4,
                     value=", ".join(r.keywords_matched)).font = data_font
            detail_row += 1

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 80
    ws2.column_dimensions['D'].width = 30

    # ---- Query metadata ----
    ws3 = wb.create_sheet("Query Info")
    info = [
        ("Query", query),
        ("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Results", len(results)),
        ("Tool", "StockValuation Company Intelligence Screener"),
    ]
    for row, (label, val) in enumerate(info, 1):
        ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=row, column=2, value=val)

    # Save
    wb.save(str(output_path))
    logger.info(f"Exported {len(results)} results to {output_path}")

    return str(output_path)
