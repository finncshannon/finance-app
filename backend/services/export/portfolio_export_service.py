import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .excel_styles import (
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, VALUE_FONT,
    HEADER_FILL, SECTION_FILL, FMT_CURRENCY, FMT_CURRENCY_DEC, FMT_PCT, FMT_NUMBER,
    apply_header_row, set_column_widths, freeze_panes, write_label_value,
)
from .pdf_styles import (
    PRIMARY, SECONDARY, ACCENT, BORDER,
    HEADER_TABLE_STYLE, fmt_currency, fmt_pct, fmt_price,
)


class PortfolioExportService:
    def __init__(self, portfolio_service=None):
        self.portfolio_svc = portfolio_service

    async def generate_excel(self, holdings=None, transactions=None, summary=None, performance=None):
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: Holdings
        ws = wb.create_sheet("Holdings")
        set_column_widths(ws, {'A': 10, 'B': 20, 'C': 12, 'D': 14, 'E': 14, 'F': 16, 'G': 16, 'H': 16, 'I': 10})
        headers = ["Ticker", "Name", "Shares", "Avg Cost", "Price", "Value", "Cost Basis", "Gain/Loss", "Weight"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
        apply_header_row(ws, 1, len(headers))

        if holdings:
            for row_idx, h in enumerate(holdings, 2):
                if isinstance(h, dict):
                    ws.cell(row=row_idx, column=1, value=h.get('ticker', '')).font = VALUE_FONT
                    ws.cell(row=row_idx, column=2, value=h.get('company_name', '')).font = LABEL_FONT
                    ws.cell(row=row_idx, column=3, value=h.get('shares', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=3).number_format = FMT_NUMBER
                    ws.cell(row=row_idx, column=4, value=h.get('avg_cost', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=4).number_format = FMT_CURRENCY_DEC
                    ws.cell(row=row_idx, column=5, value=h.get('current_price', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=5).number_format = FMT_CURRENCY_DEC
                    ws.cell(row=row_idx, column=6, value=h.get('market_value', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=6).number_format = FMT_CURRENCY
                    ws.cell(row=row_idx, column=7, value=h.get('cost_basis', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=7).number_format = FMT_CURRENCY
                    ws.cell(row=row_idx, column=8, value=h.get('gain_loss', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=8).number_format = FMT_CURRENCY
                    ws.cell(row=row_idx, column=9, value=h.get('weight', 0)).font = VALUE_FONT
                    ws.cell(row=row_idx, column=9).number_format = FMT_PCT
        freeze_panes(ws, 'B2')

        # Sheet 2: Performance
        ws2 = wb.create_sheet("Performance")
        ws2.cell(row=1, column=1, value="PERFORMANCE SUMMARY").font = HEADER_FONT
        if summary and isinstance(summary, dict):
            row = 3
            for key in ['total_value', 'total_cost', 'total_gain_loss', 'total_gain_loss_pct']:
                write_label_value(ws2, row, key.replace('_', ' ').title(), summary.get(key, 0),
                                value_fmt=FMT_PCT if 'pct' in key else FMT_CURRENCY)
                row += 1

        # Sheet 3: Allocation
        ws3 = wb.create_sheet("Allocation")
        ws3.cell(row=1, column=1, value="SECTOR ALLOCATION").font = HEADER_FONT

        # Sheet 4: Transactions
        ws4 = wb.create_sheet("Transactions")
        t_headers = ["Date", "Ticker", "Type", "Shares", "Price", "Amount"]
        for i, h in enumerate(t_headers, 1):
            c = ws4.cell(row=1, column=i, value=h)
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
        apply_header_row(ws4, 1, len(t_headers))
        set_column_widths(ws4, {'A': 14, 'B': 10, 'C': 10, 'D': 10, 'E': 12, 'F': 14})
        if transactions:
            for row_idx, txn in enumerate(transactions, 2):
                if isinstance(txn, dict):
                    ws4.cell(row=row_idx, column=1, value=txn.get('date', '')).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=2, value=txn.get('ticker', '')).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=3, value=txn.get('transaction_type', '')).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=4, value=txn.get('shares', 0)).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=5, value=txn.get('price', 0)).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=5).number_format = FMT_CURRENCY_DEC
                    ws4.cell(row=row_idx, column=6, value=txn.get('total_amount', 0)).font = VALUE_FONT
                    ws4.cell(row=row_idx, column=6).number_format = FMT_CURRENCY
        freeze_panes(ws4, 'B2')

        # Sheet 5: Dividends
        ws5 = wb.create_sheet("Dividends")
        ws5.cell(row=1, column=1, value="DIVIDEND INCOME").font = HEADER_FONT

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def generate_pdf(self, holdings=None, summary=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               topMargin=0.75*inch, bottomMargin=0.75*inch,
                               leftMargin=0.75*inch, rightMargin=0.75*inch)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('PTitle', parent=styles['Title'], fontSize=16, textColor=PRIMARY)
        heading_style = ParagraphStyle('PHeading', parent=styles['Heading2'], fontSize=12, textColor=PRIMARY)
        body_style = ParagraphStyle('PBody', parent=styles['Normal'], fontSize=9, textColor=PRIMARY)

        story = []
        story.append(Paragraph("PORTFOLIO SUMMARY", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d')}", body_style))
        story.append(Spacer(1, 18))

        # Summary stats
        if summary and isinstance(summary, dict):
            story.append(Paragraph("Overview", heading_style))
            sum_data = [
                ["Total Value", fmt_currency(summary.get('total_value', 0))],
                ["Total Cost", fmt_currency(summary.get('total_cost', 0))],
                ["Gain/Loss", fmt_currency(summary.get('total_gain_loss', 0))],
                ["Return", fmt_pct(summary.get('total_gain_loss_pct', 0))],
            ]
            t = Table(sum_data, colWidths=[2.5*inch, 2*inch])
            t.setStyle(HEADER_TABLE_STYLE)
            story.append(t)
            story.append(Spacer(1, 18))

        # Holdings table
        if holdings:
            story.append(Paragraph("Holdings", heading_style))
            h_data = [["Ticker", "Shares", "Price", "Value", "Gain/Loss"]]
            for h in holdings[:20]:  # Top 20
                if isinstance(h, dict):
                    h_data.append([
                        h.get('ticker', ''),
                        str(h.get('shares', 0)),
                        fmt_price(h.get('current_price', 0)),
                        fmt_currency(h.get('market_value', 0)),
                        fmt_currency(h.get('gain_loss', 0)),
                    ])
            t = Table(h_data, colWidths=[1*inch, 1*inch, 1.2*inch, 1.5*inch, 1.5*inch])
            t.setStyle(HEADER_TABLE_STYLE)
            story.append(t)

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
