import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from .excel_styles import (
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, VALUE_FONT, SECTION_FONT,
    HEADER_FILL, SECTION_FILL,
    FMT_CURRENCY, FMT_CURRENCY_DEC, FMT_PCT, FMT_NUMBER,
    apply_header_row, set_column_widths, freeze_panes,
)


class ResearchExportService:
    def __init__(self, db=None):
        self.db = db

    async def generate_excel(self, ticker, financials=None, ratios=None):
        wb = Workbook()
        wb.remove(wb.active)

        if financials:
            self._build_income_statement(wb, ticker, financials)
            self._build_balance_sheet(wb, ticker, financials)
            self._build_cash_flow(wb, ticker, financials)

        if ratios:
            self._build_ratios(wb, ticker, ratios)

        if not financials and not ratios:
            ws = wb.create_sheet("Info")
            ws.cell(row=1, column=1, value=f"No financial data available for {ticker}").font = LABEL_FONT

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _build_statement_sheet(self, wb, title, ticker, financials, line_items):
        """Generic Bloomberg-style statement builder."""
        ws = wb.create_sheet(title)
        years = sorted(set(f.get('fiscal_year', 0) for f in financials), reverse=True)

        # Column A = labels, B onwards = years (newest first)
        set_column_widths(ws, {'A': 28})
        ws.cell(row=1, column=1, value=f"{ticker} — {title}").font = HEADER_FONT

        # Year headers
        for i, yr in enumerate(years):
            col = i + 2
            c = ws.cell(row=2, column=col, value=str(yr))
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col)].width = 16
        apply_header_row(ws, 2, len(years) + 1)

        # Data rows
        year_data = {}
        for f in financials:
            year_data[f.get('fiscal_year', 0)] = f

        for item_idx, (label, key, fmt, is_bold) in enumerate(line_items):
            row = item_idx + 3
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = SECTION_FONT if is_bold else LABEL_FONT

            for yr_idx, yr in enumerate(years):
                col = yr_idx + 2
                f = year_data.get(yr, {})
                val = f.get(key)
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Calibri', bold=is_bold, size=10, color='1F2937')
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right')
                # Negative in red
                if isinstance(val, (int, float)) and val < 0:
                    c.font = Font(name='Calibri', bold=is_bold, size=10, color='DC2626')

        freeze_panes(ws, 'B3')

    def _build_income_statement(self, wb, ticker, financials):
        items = [
            ("Revenue", "revenue", FMT_CURRENCY, True),
            ("Cost of Revenue", "cost_of_revenue", FMT_CURRENCY, False),
            ("Gross Profit", "gross_profit", FMT_CURRENCY, True),
            ("Gross Margin", "gross_margin", FMT_PCT, False),
            ("R&D Expense", "rd_expense", FMT_CURRENCY, False),
            ("SG&A Expense", "sga_expense", FMT_CURRENCY, False),
            ("Operating Expense", "operating_expense", FMT_CURRENCY, False),
            ("EBIT", "ebit", FMT_CURRENCY, True),
            ("Operating Margin", "operating_margin", FMT_PCT, False),
            ("Interest Expense", "interest_expense", FMT_CURRENCY, False),
            ("Tax Provision", "tax_provision", FMT_CURRENCY, False),
            ("Net Income", "net_income", FMT_CURRENCY, True),
            ("Net Margin", "net_margin", FMT_PCT, False),
            ("EBITDA", "ebitda", FMT_CURRENCY, True),
            ("EBITDA Margin", "ebitda_margin", FMT_PCT, False),
            ("EPS (Basic)", "eps_basic", FMT_CURRENCY_DEC, False),
            ("EPS (Diluted)", "eps_diluted", FMT_CURRENCY_DEC, False),
        ]
        self._build_statement_sheet(wb, "Income Statement", ticker, financials, items)

    def _build_balance_sheet(self, wb, ticker, financials):
        items = [
            ("Total Assets", "total_assets", FMT_CURRENCY, True),
            ("Current Assets", "current_assets", FMT_CURRENCY, False),
            ("Cash & Equivalents", "cash_and_equivalents", FMT_CURRENCY, False),
            ("Total Liabilities", "total_liabilities", FMT_CURRENCY, True),
            ("Current Liabilities", "current_liabilities", FMT_CURRENCY, False),
            ("Long-Term Debt", "long_term_debt", FMT_CURRENCY, False),
            ("Total Debt", "total_debt", FMT_CURRENCY, False),
            ("Stockholders' Equity", "stockholders_equity", FMT_CURRENCY, True),
            ("Working Capital", "working_capital", FMT_CURRENCY, False),
            ("Net Debt", "net_debt", FMT_CURRENCY, False),
            ("Debt / Equity", "debt_to_equity", '0.00"x"', False),
        ]
        self._build_statement_sheet(wb, "Balance Sheet", ticker, financials, items)

    def _build_cash_flow(self, wb, ticker, financials):
        items = [
            ("Operating Cash Flow", "operating_cash_flow", FMT_CURRENCY, True),
            ("Capital Expenditure", "capital_expenditure", FMT_CURRENCY, False),
            ("Free Cash Flow", "free_cash_flow", FMT_CURRENCY, True),
            ("FCF Margin", "fcf_margin", FMT_PCT, False),
            ("Investing Cash Flow", "investing_cash_flow", FMT_CURRENCY, False),
            ("Financing Cash Flow", "financing_cash_flow", FMT_CURRENCY, False),
            ("Dividends Paid", "dividends_paid", FMT_CURRENCY, False),
            ("D&A", "depreciation_amortization", FMT_CURRENCY, False),
            ("Shares Outstanding", "shares_outstanding", FMT_NUMBER, False),
        ]
        self._build_statement_sheet(wb, "Cash Flow", ticker, financials, items)

    def _build_ratios(self, wb, ticker, ratios):
        ws = wb.create_sheet("Ratios")
        set_column_widths(ws, {'A': 28, 'B': 16})
        ws.cell(row=1, column=1, value=f"{ticker} — Financial Ratios").font = HEADER_FONT
        row = 3

        if not isinstance(ratios, dict):
            return

        # Group flat metrics into categories
        RATIO_GROUPS = {
            'profitability': ['gross_margin', 'operating_margin', 'net_margin', 'ebitda_margin', 'fcf_margin'],
            'returns': ['roe', 'roa', 'roic'],
            'leverage': ['debt_to_equity', 'net_debt_to_ebitda', 'interest_coverage', 'debt_to_assets'],
            'valuation': ['pe_ratio', 'pe_forward', 'price_to_book', 'price_to_sales', 'ev_to_ebitda', 'ev_to_revenue', 'fcf_yield', 'earnings_yield', 'dividend_yield'],
            'efficiency': ['asset_turnover'],
            'growth': ['revenue_growth_yoy', 'net_income_growth_yoy', 'eps_growth_yoy', 'ebitda_growth_yoy', 'fcf_growth_yoy', 'revenue_cagr_3y', 'revenue_cagr_5y', 'eps_cagr_3y', 'eps_cagr_5y'],
        }

        for group_name, metric_keys in RATIO_GROUPS.items():
            # Collect metrics that have values
            group_metrics = {k: ratios.get(k) for k in metric_keys if ratios.get(k) is not None}
            if not group_metrics:
                continue

            ws.cell(row=row, column=1, value=group_name.upper()).font = SECTION_FONT
            ws.cell(row=row, column=1).fill = SECTION_FILL
            ws.cell(row=row, column=2).fill = SECTION_FILL
            row += 1

            for metric, val in group_metrics.items():
                ws.cell(row=row, column=1, value=metric.replace('_', ' ').title()).font = LABEL_FONT
                c = ws.cell(row=row, column=2, value=val)
                c.font = VALUE_FONT
                if 'margin' in metric or 'growth' in metric or 'yield' in metric or metric in ('roe', 'roa', 'roic') or 'cagr' in metric:
                    c.number_format = FMT_PCT
                else:
                    c.number_format = '0.00'
                row += 1
            row += 1

        freeze_panes(ws, 'B3')
