import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from .excel_styles import (
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, VALUE_FONT, SECTION_FONT,
    HEADER_FILL, SECTION_FILL, INPUT_FILL, POSITIVE_FILL, NEGATIVE_FILL,
    FMT_CURRENCY, FMT_CURRENCY_DEC, FMT_PCT, FMT_MULTIPLE, FMT_NUMBER,
    THIN_BORDER, SECTION_BORDER,
    apply_header_row, set_column_widths, freeze_panes, write_label_value,
)
from .excel_formula_builder import cell_ref, growth_formula, margin_formula, pv_formula, tv_perpetuity_formula, subtract_formula, multiply_formula, divide_formula, sum_formula, same_sheet_ref


class ModelExcelService:
    def __init__(self, model_repo=None):
        self.model_repo = model_repo

    async def generate(self, ticker, model_type, assumptions, engine_result, historical_data, current_price):
        wb = Workbook()
        wb.remove(wb.active)

        # Track Assumptions sheet cell positions for cross-references
        assumption_cells = {}

        self._build_summary(wb, ticker, model_type, assumptions, engine_result, current_price)
        self._build_assumptions(wb, ticker, model_type, assumptions, assumption_cells)
        self._build_projections(wb, ticker, model_type, assumptions, engine_result, assumption_cells)
        self._build_dcf_output(wb, ticker, engine_result, assumption_cells)
        self._build_sensitivity(wb, ticker, engine_result)
        self._build_scenarios(wb, ticker, engine_result)
        self._build_comps(wb, ticker, engine_result)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _build_summary(self, wb, ticker, model_type, assumptions, engine_result, current_price):
        ws = wb.create_sheet("Summary")
        set_column_widths(ws, {'A': 30, 'B': 20, 'C': 20, 'D': 20})

        # Title
        ws.cell(row=1, column=1, value="VALUATION REPORT").font = HEADER_FONT
        ws.cell(row=2, column=1, value=f"{ticker} — {model_type.upper()} Model").font = SUBHEADER_FONT
        ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = LABEL_FONT

        # Key values
        row = 5
        base = engine_result.get('scenarios', {}).get('base', {})
        implied = base.get('implied_price', engine_result.get('weighted_implied_price', 0))
        upside = ((implied / current_price) - 1) * 100 if current_price else 0

        write_label_value(ws, row, "Current Price", current_price, value_fmt=FMT_CURRENCY_DEC)
        row += 1
        write_label_value(ws, row, "Intrinsic Value (Base)", implied, value_fmt=FMT_CURRENCY_DEC)
        row += 1
        val_cell = write_label_value(ws, row, "Upside / Downside", upside / 100, value_fmt=FMT_PCT)
        val_cell.fill = POSITIVE_FILL if upside > 0 else NEGATIVE_FILL
        row += 2

        # Scenario summary
        ws.cell(row=row, column=1, value="SCENARIO SUMMARY").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        ws.cell(row=row, column=3).fill = SECTION_FILL
        row += 1
        headers = ["Scenario", "Implied Price", "Weight"]
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
        row += 1

        for name in ['bear', 'base', 'bull']:
            sc = engine_result.get('scenarios', {}).get(name, {})
            if not sc:
                continue
            ws.cell(row=row, column=1, value=name.capitalize()).font = LABEL_FONT
            ws.cell(row=row, column=2, value=sc.get('implied_price', 0)).font = VALUE_FONT
            ws.cell(row=row, column=2).number_format = FMT_CURRENCY_DEC
            ws.cell(row=row, column=3, value=sc.get('scenario_weight', 0)).font = VALUE_FONT
            ws.cell(row=row, column=3).number_format = FMT_PCT
            row += 1

        row += 1
        # Key assumptions summary
        ws.cell(row=row, column=1, value="KEY ASSUMPTIONS").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        row += 1
        dcf_a = assumptions.get('model_assumptions', {}).get('dcf', {})
        key_items = [
            ("WACC", dcf_a.get('wacc', 0), FMT_PCT),
            ("Terminal Growth", dcf_a.get('terminal_growth_rate', 0), FMT_PCT),
            ("Tax Rate", dcf_a.get('tax_rate', 0), FMT_PCT),
            ("Projection Years", dcf_a.get('projection_years', 10), FMT_NUMBER),
        ]
        for label, val, fmt in key_items:
            write_label_value(ws, row, label, val, value_fmt=fmt)
            row += 1

        freeze_panes(ws, 'A5')

    def _build_assumptions(self, wb, ticker, model_type, assumptions, assumption_cells):
        ws = wb.create_sheet("Assumptions")
        set_column_widths(ws, {'A': 30, 'B': 15, 'C': 10, 'D': 15})

        ws.cell(row=1, column=1, value="ASSUMPTIONS").font = HEADER_FONT
        ws.cell(row=1, column=1).fill = HEADER_FILL
        apply_header_row(ws, 1, 4)

        dcf_a = assumptions.get('model_assumptions', {}).get('dcf', {})
        row = 3

        # Key parameters (named ranges tracked)
        params = [
            ("WACC", dcf_a.get('wacc', 0.10), FMT_PCT, "%", "wacc"),
            ("Terminal Growth Rate", dcf_a.get('terminal_growth_rate', 0.025), FMT_PCT, "%", "terminal_growth"),
            ("Tax Rate", dcf_a.get('tax_rate', 0.21), FMT_PCT, "%", "tax_rate"),
            ("Projection Years", dcf_a.get('projection_years', 10), FMT_NUMBER, "years", "proj_years"),
            ("CapEx % of Revenue", dcf_a.get('capex_to_revenue', 0.05), FMT_PCT, "%", "capex_pct"),
            ("NWC % of Revenue", dcf_a.get('nwc_pct_of_revenue', 0.05), FMT_PCT, "%", "nwc_pct"),
            ("D&A % of Revenue", dcf_a.get('da_pct_of_revenue', 0.04), FMT_PCT, "%", "da_pct"),
        ]

        ws.cell(row=row, column=1, value="KEY PARAMETERS").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        for label, val, fmt, unit, key in params:
            val_cell = write_label_value(ws, row, label, val, value_fmt=fmt, is_input=True)
            ws.cell(row=row, column=3, value=unit).font = LABEL_FONT
            ws.cell(row=row, column=4, value="Engine").font = LABEL_FONT
            assumption_cells[key] = cell_ref("Assumptions", 2, row)
            row += 1

        row += 1
        # Shares outstanding + net debt
        shares = assumptions.get('model_assumptions', {}).get('dcf', {}).get('shares_outstanding', 0)
        net_debt_val = assumptions.get('model_assumptions', {}).get('dcf', {}).get('net_debt', 0)

        ws.cell(row=row, column=1, value="EQUITY BRIDGE").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        val_cell = write_label_value(ws, row, "Shares Outstanding", shares, value_fmt=FMT_NUMBER, is_input=True)
        assumption_cells['shares'] = cell_ref("Assumptions", 2, row)
        row += 1
        val_cell = write_label_value(ws, row, "Net Debt", net_debt_val, value_fmt=FMT_CURRENCY, is_input=True)
        assumption_cells['net_debt'] = cell_ref("Assumptions", 2, row)
        row += 1

        row += 1
        # Revenue growth rates by year
        ws.cell(row=row, column=1, value="REVENUE GROWTH RATES").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        growth_rates = dcf_a.get('revenue_growth_rates', [])
        for i, rate in enumerate(growth_rates):
            yr_label = f"Year {i + 1}"
            val_cell = write_label_value(ws, row, yr_label, rate, value_fmt=FMT_PCT, is_input=True)
            assumption_cells[f'growth_yr{i+1}'] = cell_ref("Assumptions", 2, row)
            row += 1

        row += 1
        # Operating margins by year
        ws.cell(row=row, column=1, value="OPERATING MARGINS").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        for c in range(1, 5):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        margins = dcf_a.get('operating_margins', [])
        for i, margin in enumerate(margins):
            yr_label = f"Year {i + 1}"
            val_cell = write_label_value(ws, row, yr_label, margin, value_fmt=FMT_PCT, is_input=True)
            assumption_cells[f'op_margin_yr{i+1}'] = cell_ref("Assumptions", 2, row)
            row += 1

        freeze_panes(ws, 'B2')

    def _build_projections(self, wb, ticker, model_type, assumptions, engine_result, assumption_cells):
        ws = wb.create_sheet("Projections")

        # Get base scenario projection table
        base_scenario = engine_result.get('scenarios', {}).get('base', {})
        proj_table = base_scenario.get('projection_table', [])
        n_years = len(proj_table)
        if n_years == 0:
            ws.cell(row=1, column=1, value="No projection data available").font = LABEL_FONT
            return

        # Column layout: A=labels, B onwards = years
        set_column_widths(ws, {'A': 25})
        for i in range(n_years):
            col_letter = get_column_letter(i + 2)
            ws.column_dimensions[col_letter].width = 16

        # Header row with years
        ws.cell(row=1, column=1, value="PROJECTIONS").font = HEADER_FONT
        for i, yr_data in enumerate(proj_table):
            col = i + 2
            c = ws.cell(row=2, column=col, value=f"Year {yr_data.get('year', i+1)}")
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')
        apply_header_row(ws, 2, n_years + 1)

        # Line items — USE LIVE FORMULAS where possible
        line_items = [
            ("Revenue", "revenue", FMT_CURRENCY),
            ("Revenue Growth", "revenue_growth", FMT_PCT),
            ("COGS", "cogs", FMT_CURRENCY),
            ("Gross Profit", "gross_profit", FMT_CURRENCY),
            ("Gross Margin", "gross_margin", FMT_PCT),
            ("Operating Expenses", "opex", FMT_CURRENCY),
            ("EBIT", "ebit", FMT_CURRENCY),
            ("Operating Margin", "operating_margin", FMT_PCT),
            ("D&A", "da", FMT_CURRENCY),
            ("EBITDA", "ebitda", FMT_CURRENCY),
            ("EBITDA Margin", "ebitda_margin", FMT_PCT),
            ("Taxes", "taxes", FMT_CURRENCY),
            ("NOPAT", "nopat", FMT_CURRENCY),
            ("CapEx", "capex", FMT_CURRENCY),
            ("NWC Change", "nwc_change", FMT_CURRENCY),
            ("Free Cash Flow", "fcf", FMT_CURRENCY),
            ("FCF Margin", "fcf_margin", FMT_PCT),
            ("Discount Factor", "discount_factor", '0.0000'),
            ("PV of FCF", "pv_fcf", FMT_CURRENCY),
        ]

        wacc_ref = assumption_cells.get('wacc', '')
        tax_ref = assumption_cells.get('tax_rate', '')

        for item_idx, (label, key, fmt) in enumerate(line_items):
            row = item_idx + 3
            # Label cell
            lbl = ws.cell(row=row, column=1, value=label)
            lbl.font = LABEL_FONT
            if key in ('gross_profit', 'ebit', 'ebitda', 'nopat', 'fcf'):
                lbl.font = SECTION_FONT

            for yr_idx, yr_data in enumerate(proj_table):
                col = yr_idx + 2
                val = yr_data.get(key, 0)

                # For key computed items, write LIVE FORMULAS
                if key == 'revenue' and yr_idx > 0 and assumption_cells.get(f'growth_yr{yr_idx+1}'):
                    prev_rev = same_sheet_ref(col - 1, row)
                    growth_ref = assumption_cells[f'growth_yr{yr_idx+1}']
                    ws.cell(row=row, column=col, value=growth_formula(prev_rev, growth_ref))
                elif key == 'ebit' and assumption_cells.get(f'op_margin_yr{yr_idx+1}'):
                    rev_row = 3  # Revenue is row 3
                    rev_ref = same_sheet_ref(col, rev_row)
                    margin_ref = assumption_cells[f'op_margin_yr{yr_idx+1}']
                    ws.cell(row=row, column=col, value=margin_formula(rev_ref, margin_ref))
                elif key == 'taxes' and tax_ref:
                    ebit_row = 3 + [li[1] for li in line_items].index('ebit')
                    ebit_ref = same_sheet_ref(col, ebit_row)
                    ws.cell(row=row, column=col, value=multiply_formula(ebit_ref, tax_ref))
                elif key == 'nopat':
                    ebit_row = 3 + [li[1] for li in line_items].index('ebit')
                    tax_row = 3 + [li[1] for li in line_items].index('taxes')
                    ebit_ref = same_sheet_ref(col, ebit_row)
                    taxes_ref = same_sheet_ref(col, tax_row)
                    ws.cell(row=row, column=col, value=subtract_formula(ebit_ref, taxes_ref))
                elif key == 'gross_profit':
                    rev_row = 3
                    cogs_row = 3 + [li[1] for li in line_items].index('cogs')
                    ws.cell(row=row, column=col, value=subtract_formula(same_sheet_ref(col, rev_row), same_sheet_ref(col, cogs_row)))
                elif key == 'ebitda':
                    ebit_row = 3 + [li[1] for li in line_items].index('ebit')
                    da_row = 3 + [li[1] for li in line_items].index('da')
                    ws.cell(row=row, column=col, value=f"={same_sheet_ref(col, ebit_row)}+{same_sheet_ref(col, da_row)}")
                elif key == 'fcf':
                    nopat_row = 3 + [li[1] for li in line_items].index('nopat')
                    da_row = 3 + [li[1] for li in line_items].index('da')
                    capex_row = 3 + [li[1] for li in line_items].index('capex')
                    nwc_row = 3 + [li[1] for li in line_items].index('nwc_change')
                    ws.cell(row=row, column=col, value=f"={same_sheet_ref(col, nopat_row)}+{same_sheet_ref(col, da_row)}-{same_sheet_ref(col, capex_row)}-{same_sheet_ref(col, nwc_row)}")
                elif key == 'pv_fcf' and wacc_ref:
                    fcf_row = 3 + [li[1] for li in line_items].index('fcf')
                    fcf_ref = same_sheet_ref(col, fcf_row)
                    ws.cell(row=row, column=col, value=pv_formula(fcf_ref, wacc_ref, yr_idx + 1))
                elif key == 'discount_factor' and wacc_ref:
                    ws.cell(row=row, column=col, value=f"=1/((1+{wacc_ref})^{yr_idx+1})")
                else:
                    # Static value fallback
                    ws.cell(row=row, column=col, value=val)

                ws.cell(row=row, column=col).font = VALUE_FONT
                ws.cell(row=row, column=col).number_format = fmt
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')

        # Section borders
        for key in ('gross_profit', 'ebit', 'ebitda', 'nopat', 'fcf'):
            idx = [li[1] for li in line_items].index(key)
            for col in range(1, n_years + 2):
                ws.cell(row=3 + idx, column=col).border = THIN_BORDER

        freeze_panes(ws, 'B3')

    def _build_dcf_output(self, wb, ticker, engine_result, assumption_cells):
        ws = wb.create_sheet("DCF Output")
        set_column_widths(ws, {'A': 30, 'B': 20})

        ws.cell(row=1, column=1, value="DCF VALUATION").font = HEADER_FONT
        row = 3

        # Waterfall
        base = engine_result.get('scenarios', {}).get('base', {})
        waterfall = engine_result.get('waterfall', {}).get('steps', [])

        ws.cell(row=row, column=1, value="VALUATION WATERFALL").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2).fill = SECTION_FILL
        row += 1

        for step in waterfall:
            lbl = ws.cell(row=row, column=1, value=step.get('label', ''))
            lbl.font = LABEL_FONT
            if step.get('step_type') in ('subtotal', 'end'):
                lbl.font = SECTION_FONT
            val = ws.cell(row=row, column=2, value=step.get('value', 0))
            val.font = VALUE_FONT
            val.number_format = FMT_CURRENCY
            row += 1

        if not waterfall:
            # Build from scenario data
            items = [
                ("PV of FCFs", base.get('pv_fcf_total', 0)),
                ("PV of Terminal Value", base.get('pv_terminal_value', 0)),
                ("Enterprise Value", base.get('enterprise_value', 0)),
                ("Less: Net Debt", -(assumption_cells.get('net_debt', 0) if isinstance(assumption_cells, dict) else 0)),
                ("Equity Value", base.get('equity_value', 0)),
                ("Implied Price/Share", base.get('implied_price', 0)),
            ]
            for label, val in items:
                write_label_value(ws, row, label, val, value_fmt=FMT_CURRENCY)
                row += 1

        row += 1
        ws.cell(row=row, column=1, value="TV % of Enterprise Value").font = LABEL_FONT
        ws.cell(row=row, column=2, value=base.get('tv_pct_of_ev', 0)).font = VALUE_FONT
        ws.cell(row=row, column=2).number_format = FMT_PCT

        freeze_panes(ws, 'A3')

    def _build_sensitivity(self, wb, ticker, engine_result):
        ws = wb.create_sheet("Sensitivity")
        set_column_widths(ws, {'A': 20})

        ws.cell(row=1, column=1, value="SENSITIVITY ANALYSIS").font = HEADER_FONT
        ws.cell(row=3, column=1, value="WACC vs Terminal Growth → Implied Price").font = SUBHEADER_FONT

        base = engine_result.get('scenarios', {}).get('base', {})
        base_wacc = base.get('wacc', 0.10)
        base_tg = base.get('terminal_growth_rate', 0.025)
        base_price = base.get('implied_price', 0)

        # Generate grid
        wacc_range = [base_wacc + (i - 4) * 0.005 for i in range(9)]
        tg_range = [base_tg + (i - 3) * 0.005 for i in range(7)]

        # Column headers (terminal growth)
        row = 5
        ws.cell(row=row, column=1, value="WACC \\ TG").font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        for j, tg in enumerate(tg_range):
            c = ws.cell(row=row, column=j + 2, value=tg)
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
            c.number_format = FMT_PCT
            c.alignment = Alignment(horizontal='center')

        # Row data
        for i, wacc in enumerate(wacc_range):
            r = row + 1 + i
            wc = ws.cell(row=r, column=1, value=wacc)
            wc.font = VALUE_FONT
            wc.number_format = FMT_PCT
            wc.fill = HEADER_FILL

            for j, tg in enumerate(tg_range):
                # Approximate: scale base price by ratio of discount rates
                if wacc > tg:
                    approx = base_price * (base_wacc - base_tg) / (wacc - tg) if (base_wacc - base_tg) > 0 else base_price
                else:
                    approx = 0
                c = ws.cell(row=r, column=j + 2, value=round(approx, 2))
                c.font = VALUE_FONT
                c.number_format = FMT_CURRENCY_DEC
                c.alignment = Alignment(horizontal='right')
                # Highlight base case
                if abs(wacc - base_wacc) < 0.001 and abs(tg - base_tg) < 0.001:
                    c.fill = INPUT_FILL
                    c.font = Font(name='Calibri', bold=True, size=10)

        freeze_panes(ws, 'B6')

    def _build_scenarios(self, wb, ticker, engine_result):
        ws = wb.create_sheet("Scenarios")
        set_column_widths(ws, {'A': 25, 'B': 18, 'C': 18, 'D': 18})

        ws.cell(row=1, column=1, value="SCENARIO COMPARISON").font = HEADER_FONT

        scenarios = engine_result.get('scenarios', {})
        names = ['bear', 'base', 'bull']
        available = [n for n in names if n in scenarios]

        if not available:
            ws.cell(row=3, column=1, value="No scenario data available").font = LABEL_FONT
            return

        # Headers
        row = 3
        ws.cell(row=row, column=1, value="Metric").font = SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        for i, name in enumerate(available):
            c = ws.cell(row=row, column=i + 2, value=name.capitalize())
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')
        apply_header_row(ws, row, len(available) + 1)
        row += 1

        # Metrics
        metrics = [
            ("Implied Price", "implied_price", FMT_CURRENCY_DEC),
            ("Enterprise Value", "enterprise_value", FMT_CURRENCY),
            ("WACC", "wacc", FMT_PCT),
            ("Terminal Growth", "terminal_growth_rate", FMT_PCT),
            ("Weight", "scenario_weight", FMT_PCT),
            ("PV of FCFs", "pv_fcf_total", FMT_CURRENCY),
            ("PV Terminal Value", "pv_terminal_value", FMT_CURRENCY),
            ("TV % of EV", "tv_pct_of_ev", FMT_PCT),
            ("Upside/Downside", "upside_downside_pct", FMT_PCT),
        ]

        for label, key, fmt in metrics:
            ws.cell(row=row, column=1, value=label).font = LABEL_FONT
            for i, name in enumerate(available):
                sc = scenarios[name]
                val = sc.get(key, 0)
                if val is None:
                    val = 0
                c = ws.cell(row=row, column=i + 2, value=val)
                c.font = VALUE_FONT
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right')
            row += 1

        # Weighted composite
        row += 1
        ws.cell(row=row, column=1, value="Weighted Implied Price").font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=2, value=engine_result.get('weighted_implied_price', 0)).font = Font(name='Calibri', bold=True, size=11)
        ws.cell(row=row, column=2).number_format = FMT_CURRENCY_DEC

        freeze_panes(ws, 'B4')

    def _build_comps(self, wb, ticker, engine_result):
        ws = wb.create_sheet("Comps Table")
        set_column_widths(ws, {'A': 10, 'B': 25, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 12, 'H': 12, 'I': 12})

        ws.cell(row=1, column=1, value="COMPARABLE COMPANIES").font = HEADER_FONT

        peer_group = engine_result.get('peer_group', {})
        peers = peer_group.get('peers', []) if isinstance(peer_group, dict) else []

        if not peers:
            ws.cell(row=3, column=1, value="No comparable company data available").font = LABEL_FONT
            return

        # Headers
        row = 3
        headers = ["Ticker", "Company", "Market Cap", "EV", "Revenue", "EBITDA", "P/E", "EV/EBITDA", "EV/Rev"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=i + 1, value=h)
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
        apply_header_row(ws, row, len(headers))
        row += 1

        for peer in peers:
            if isinstance(peer, dict):
                ws.cell(row=row, column=1, value=peer.get('ticker', '')).font = VALUE_FONT
                ws.cell(row=row, column=2, value=peer.get('company_name', '')).font = LABEL_FONT
                ws.cell(row=row, column=3, value=peer.get('market_cap', 0)).font = VALUE_FONT
                ws.cell(row=row, column=3).number_format = FMT_CURRENCY
                ws.cell(row=row, column=4, value=peer.get('enterprise_value', 0)).font = VALUE_FONT
                ws.cell(row=row, column=4).number_format = FMT_CURRENCY
                ws.cell(row=row, column=5, value=peer.get('revenue', 0)).font = VALUE_FONT
                ws.cell(row=row, column=5).number_format = FMT_CURRENCY
                ws.cell(row=row, column=6, value=peer.get('ebitda', 0)).font = VALUE_FONT
                ws.cell(row=row, column=6).number_format = FMT_CURRENCY
                ws.cell(row=row, column=7, value=peer.get('pe', 0)).font = VALUE_FONT
                ws.cell(row=row, column=7).number_format = FMT_MULTIPLE
                ws.cell(row=row, column=8, value=peer.get('ev_ebitda', 0)).font = VALUE_FONT
                ws.cell(row=row, column=8).number_format = FMT_MULTIPLE
                ws.cell(row=row, column=9, value=peer.get('ev_revenue', 0)).font = VALUE_FONT
                ws.cell(row=row, column=9).number_format = FMT_MULTIPLE
                row += 1

        # Aggregate stats
        agg = engine_result.get('aggregate_multiples', {})
        if agg:
            row += 1
            ws.cell(row=row, column=1).border = SECTION_BORDER
            for stat_name in ['median', 'mean']:
                ws.cell(row=row, column=2, value=stat_name.capitalize()).font = SECTION_FONT
                for mult_name, col_idx in [('pe', 7), ('ev_ebitda', 8), ('ev_revenue', 9)]:
                    stats = agg.get(mult_name, {})
                    if isinstance(stats, dict):
                        ws.cell(row=row, column=col_idx, value=stats.get(stat_name, 0)).font = VALUE_FONT
                        ws.cell(row=row, column=col_idx).number_format = FMT_MULTIPLE
                row += 1

        freeze_panes(ws, 'C4')
