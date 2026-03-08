from openpyxl.utils import get_column_letter

def cell_ref(sheet_name, col, row):
    """Build cross-sheet cell reference like 'Assumptions'!B3"""
    col_letter = get_column_letter(col)
    return f"'{sheet_name}'!{col_letter}{row}"

def same_sheet_ref(col, row):
    col_letter = get_column_letter(col)
    return f"{col_letter}{row}"

def growth_formula(prev_cell, growth_cell):
    """= prev_cell * (1 + growth_cell)"""
    return f"={prev_cell}*(1+{growth_cell})"

def margin_formula(revenue_cell, margin_cell):
    """= revenue_cell * margin_cell"""
    return f"={revenue_cell}*{margin_cell}"

def subtract_formula(a, b):
    return f"={a}-{b}"

def multiply_formula(a, b):
    return f"={a}*{b}"

def divide_formula(a, b):
    return f"={a}/{b}"

def sum_formula(cells):
    return f"=SUM({','.join(cells)})"

def pv_formula(cf_cell, wacc_ref, year):
    """= cf_cell / (1 + wacc_ref) ^ year"""
    return f"={cf_cell}/((1+{wacc_ref})^{year})"

def tv_perpetuity_formula(terminal_fcf, tg_ref, wacc_ref):
    """= terminal_fcf * (1 + tg_ref) / (wacc_ref - tg_ref)"""
    return f"={terminal_fcf}*(1+{tg_ref})/({wacc_ref}-{tg_ref})"
