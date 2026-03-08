from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# Fonts
HEADER_FONT = Font(name='Calibri', bold=True, size=14, color='1F2937')
SUBHEADER_FONT = Font(name='Calibri', bold=True, size=11, color='374151')
LABEL_FONT = Font(name='Calibri', size=10, color='4B5563')
VALUE_FONT = Font(name='Calibri', size=10, color='1F2937')
SECTION_FONT = Font(name='Calibri', bold=True, size=10, color='1F2937')

# Fills
HEADER_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
POSITIVE_FILL = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
NEGATIVE_FILL = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

# Number formats
FMT_CURRENCY = '#,##0'
FMT_CURRENCY_DEC = '#,##0.00'
FMT_PCT = '0.0%'
FMT_MULTIPLE = '0.0"x"'
FMT_NUMBER = '#,##0'

# Borders
THIN_BORDER = Border(bottom=Side(style='thin', color='D1D5DB'))
SECTION_BORDER = Border(bottom=Side(style='medium', color='9CA3AF'))

def apply_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = SECTION_BORDER

def set_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def freeze_panes(ws, cell='B2'):
    ws.freeze_panes = cell

def write_label_value(ws, row, label, value, col_label=1, col_value=2, value_fmt=None, is_input=False):
    lbl_cell = ws.cell(row=row, column=col_label, value=label)
    lbl_cell.font = LABEL_FONT
    val_cell = ws.cell(row=row, column=col_value, value=value)
    val_cell.font = VALUE_FONT
    if value_fmt:
        val_cell.number_format = value_fmt
    if is_input:
        val_cell.fill = INPUT_FILL
    return val_cell
