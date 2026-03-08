import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .excel_styles import (
    HEADER_FONT, SUBHEADER_FONT, LABEL_FONT, VALUE_FONT,
    HEADER_FILL, FMT_CURRENCY, FMT_PCT, FMT_NUMBER,
    apply_header_row, set_column_widths, freeze_panes,
)


class ScannerExportService:
    async def generate_excel(self, results, screen_config=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        if not results:
            ws.cell(row=1, column=1, value="No results").font = LABEL_FONT
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.getvalue()

        # Determine columns from first result
        cols = list(results[0].keys())

        # Header row
        for i, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=i, value=col.replace('_', ' ').title())
            c.font = SUBHEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal='center')
        apply_header_row(ws, 1, len(cols))

        # Data rows
        for row_idx, result in enumerate(results, 2):
            for col_idx, key in enumerate(cols, 1):
                val = result.get(key)
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.font = VALUE_FONT
                if isinstance(val, float):
                    if 'margin' in key or 'growth' in key or 'pct' in key or 'yield' in key:
                        c.number_format = FMT_PCT
                    elif 'cap' in key or 'revenue' in key or 'income' in key or 'value' in key:
                        c.number_format = FMT_CURRENCY
                    else:
                        c.number_format = '0.00'

        # Auto-filter and freeze
        ws.auto_filter.ref = f"A1:{chr(64+len(cols))}{len(results)+1}"
        freeze_panes(ws, 'B2')

        # Auto-width columns
        for i, col in enumerate(cols, 1):
            from openpyxl.utils import get_column_letter
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = max(12, len(col) + 4)

        # Sheet 2: Screen Config
        if screen_config:
            ws2 = wb.create_sheet("Screen Config")
            ws2.cell(row=1, column=1, value="SCREEN CONFIGURATION").font = HEADER_FONT
            ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = LABEL_FONT
            row = 4
            if isinstance(screen_config, dict):
                for key, val in screen_config.items():
                    ws2.cell(row=row, column=1, value=str(key)).font = LABEL_FONT
                    ws2.cell(row=row, column=2, value=str(val)).font = VALUE_FONT
                    row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def generate_csv(self, results):
        if not results:
            return b""
        output = io.StringIO()
        cols = list(results[0].keys())
        writer = csv.DictWriter(output, fieldnames=cols)
        writer.writeheader()
        for row in results:
            writer.writerow(row)
        return output.getvalue().encode('utf-8')
